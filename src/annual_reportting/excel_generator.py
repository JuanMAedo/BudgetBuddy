import os

from openpyxl import Workbook, load_workbook

from utils.functionalities import day_to_name, month_to_name, year_from_excel_date

from .months import create_monthly_column_sheets

# Get paths from environment variables
monthly_income_expense_report = os.getenv("MONTHLY_REPORT_NAME")
table_start_column = os.getenv("START_COLUMN")
table_finish_column = os.getenv("FINISH_COLUMN")


def generate_excel_report(expense_income_read):
    excel_date = expense_income_read.active[table_start_column].value
    year = year_from_excel_date(excel_date)

    file_name = f"{monthly_income_expense_report} {year}.xlsx"
    if os.path.exists(file_name):
        inc_exp_excel = load_workbook(file_name)
    else:
        inc_exp_excel = Workbook()
        inc_exp_excel.create_sheet("DASHBOARD")
        inc_exp_excel.create_sheet("CATEGORY")
        for month_number, month_name in month_to_name.items():
            month_sheet = inc_exp_excel.create_sheet(month_name)
            inc_exp_excel.active = month_sheet
            create_monthly_column_sheets(month_sheet, "INCOME", "B10", "Week", "24D124", "D78740")
            create_monthly_column_sheets(month_sheet, "EXPENSE", "B2", "Week", "C06FCA", "D78740")
        inc_exp_excel.remove(inc_exp_excel["Sheet"])

    inc_exp_excel.save(file_name)
