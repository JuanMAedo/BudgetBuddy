from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string

from utils.excel_styles import thin_border
from utils.functionalities import day_to_name


def create_monthly_column_sheets(
    month_sheet, title, init_column, text_column, color_title, color_column
):
    # PARSER INIT_COLUMN TO ITERATIONS
    column_letter = init_column[0]
    column_number = column_index_from_string(column_letter)
    row_number = int(init_column[1:])

    # VERTICAL AXIS
    month_sheet[init_column] = f"{title}"
    cell_B2 = month_sheet[init_column]
    cell_B2.alignment = Alignment(horizontal="center")
    cell_B2.fill = PatternFill(
        start_color=color_title, end_color=color_title, fill_type="solid"
    )  # Color morado claro
    cell_B2.font = Font(bold=True, name="Calibri")

    for i in range(5):
        current_row = row_number + i
        cell = month_sheet.cell(row=current_row + 1, column=column_number)
        week_text = f"{text_column} {i + 1}"

        cell.value = week_text
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color=color_column, end_color=color_column, fill_type="solid"
        )  # Color naranja apagado
        cell.font = Font(bold=True, name="Calibri")

    # HORIZONTAL AXIS
    for day_number, day_name in day_to_name.items():
        current_column = column_number
        cell = month_sheet.cell(row=row_number, column=current_column + 1 + day_number)
        day_text = f"{day_name}"
        cell.value = day_text
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color="528EF6", end_color="528EF6", fill_type="solid"
        )  # Color azul pastel
        cell.font = Font(bold=True, name="Calibri")

        cell_range = month_sheet[init_column:"I7"]  # Por ejemplo, del A1 al D5
    # Aplicar el borde al rango de celdas seleccionado
    for row in cell_range:
        for cell in row:
            cell.border = thin_border
        # Aplicar el borde grueso a los lados exteriores del rango

    for row_idx, row in enumerate(cell_range, 1):
        for col_idx, cell in enumerate(row, 1):
            if row_idx == 1:  # Borde superior
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thick"),
                    bottom=Side(style="thin"),
                )
            if row_idx == len(cell_range):  # Borde inferior
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thick"),
                )
            if col_idx == 1:  # Borde izquierdo
                cell.border = Border(
                    left=Side(style="thick"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
            if col_idx == len(row):  # Borde derecho
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thick"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
